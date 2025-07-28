import streamlit as st
import pandas as pd
import io
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# --------------------------------------------------------------------------
# 함수 정의
# --------------------------------------------------------------------------

def to_excel_formatted(df, format_type=None):
    """데이터프레임을 서식이 적용된 엑셀 파일 형식의 BytesIO 객체로 변환하는 함수"""
    output = io.BytesIO()
    df_to_save = df.fillna('')
    
    if format_type == 'ecount_upload':
        df_to_save = df_to_save.rename(columns={'적요_전표': '적요', '적요_품목': '적요.1'})

    df_to_save.to_excel(output, index=False, sheet_name='Sheet1')
    
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active

    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    pink_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    if format_type == 'packing_list':
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border
        
        bundle_start_row = 2
        for row_num in range(2, sheet.max_row + 2):
            current_bundle_cell = sheet.cell(row=row_num, column=1) if row_num <= sheet.max_row else None
            
            if (current_bundle_cell and current_bundle_cell.value) or (row_num > sheet.max_row):
                if row_num > 2:
                    bundle_end_row = row_num - 1
                    prev_bundle_num_str = str(sheet.cell(row=bundle_start_row, column=1).value)
                    
                    if prev_bundle_num_str.isdigit():
                        prev_bundle_num = int(prev_bundle_num_str)
                        if prev_bundle_num % 2 != 0:
                            for r in range(bundle_start_row, bundle_end_row + 1):
                                for c in range(1, sheet.max_column + 1):
                                    sheet.cell(row=r, column=c).fill = pink_fill
                    
                    if bundle_start_row < bundle_end_row:
                        sheet.merge_cells(start_row=bundle_start_row, start_column=1, end_row=bundle_end_row, end_column=1)
                        sheet.merge_cells(start_row=bundle_start_row, start_column=4, end_row=bundle_end_row, end_column=4)
                
                bundle_start_row = row_num

    if format_type == 'quantity_summary':
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
            for cell in row:
                cell.border = thin_border
            if row_idx > 0 and row_idx % 2 != 0:
                for cell in row:
                    cell.fill = pink_fill
    
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    
    return final_output.getvalue()

@st.cache_data
def load_local_master_data(file_path="master_data.csv"):
    df_master = pd.read_csv(file_path)
    df_master = df_master.drop_duplicates(subset=['SKU코드'], keep='first')
    return df_master

def process_all_files(file1, file2, file3, df_master):
    try:
        df_smartstore = pd.read_excel(file1)
        df_ecount_orig = pd.read_excel(file2)
        df_godomall = pd.read_excel(file3)

        df_ecount_orig['original_order'] = range(len(df_ecount_orig))
        
        if '회 할인 금액' in df_godomall.columns and '회원 할인 금액' not in df_godomall.columns:
            df_godomall.rename(columns={'회 할인 금액': '회원 할인 금액'}, inplace=True)
        
        cols_to_numeric = ['상품별 품목금액', '총 배송 금액', '회원 할인 금액', '쿠폰 할인 금액', '사용된 마일리지', '총 결제 금액']
        for col in cols_to_numeric:
            if col in df_godomall.columns:
                df_godomall[col] = pd.to_numeric(df_godomall[col].astype(str).str.replace('[원,]', '', regex=True), errors='coerce').fillna(0)
        
        df_godomall['보정된_배송비'] = np.where(df_godomall.duplicated(subset=['수취인 이름', '총 결제 금액']), 0, df_godomall['총 배송 금액'])
        df_godomall['수정될_금액_고도몰'] = (df_godomall['상품별 품목금액'] + df_godomall['보정된_배송비'] - df_godomall['회원 할인 금액'] - df_godomall['쿠폰 할인 금액'] - df_godomall['사용된 마일리지'])
        
        godomall_warnings = []
        grouped_godomall = df_godomall.groupby(['수취인 이름', '총 결제 금액'])
        for (name, total_payment), group in grouped_godomall:
            calculated_total = group['수정될_금액_고도몰'].sum()
            actual_total = group['총 결제 금액'].iloc[0]
            discrepancy = calculated_total - actual_total
            if abs(discrepancy) > 1:
                warning_msg = f"- [고도몰 금액 불일치] **{name}**님의 주문(결제액:{actual_total:,.0f}원)의 계산된 금액과 실제 결제 금액이 **{discrepancy:,.0f}원** 만큼 차이납니다."
                godomall_warnings.append(warning_msg)

        df_final = df_ecount_orig.copy().rename(columns={'금액': '실결제금액'})
        
        # ▼▼▼ [핵심 수정] 컬럼명 오류를 방지하고 명확한 에러를 안내하는 로직 ▼▼▼
        
        # 1. 컬럼명의 앞뒤 공백 제거
        df_smartstore.columns = df_smartstore.columns.str.strip()
        df_godomall.columns = df_godomall.columns.str.strip()
        df_final.columns = df_final.columns.str.strip()

        # 2. 파일별로 사용할 '상품명' 컬럼의 실제 이름 정의
        ecount_name_col = 'SKU상품명'
        godo_name_col = '상품명'
        # [수정] 스마트스토어의 표준 컬럼명인 '상품명'으로 변경
        smartstore_name_col = '상품명' 

        # 3. [오류 방지] 각 파일에 필요한 핵심 컬럼이 있는지 확인
        required_cols = {
            '스마트스토어': [smartstore_name_col, '재고관리코드', '수령자명', '실결제금액'],
            '고도몰': [godo_name_col, '자체옵션코드', '수취인 이름', '상품별 품목금액'],
            '이카운트': [ecount_name_col, '재고관리코드', '수령자명', '실결제금액']
        }
        
        # 파일 이름과 데이터프레임을 매핑
        file_map = {'스마트스토어': df_smartstore, '고도몰': df_godomall, '이카운트': df_final}

        for file_name, cols in required_cols.items():
            for col in cols:
                if col not in file_map[file_name].columns:
                    st.error(f"처리 중지: '{file_name}' 파일에 필수 컬럼인 '{col}'이 없습니다.")
                    st.info(f"업로드하신 '{file_name}' 파일의 실제 열(컬럼) 이름을 확인하고 코드의 변수 값을 수정해주세요.")
                    return None, None, None, None, False, f"{file_name} 파일 컬럼 오류", []

        # 4. 데이터 값의 공백 제거 및 타입 통일
        for df in [df_final, df_smartstore, df_godomall]:
             for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).str.strip().replace('nan', '')
        
        # 5. 각 데이터프레임에 '최종키'와 '순번' 생성
        df_godomall['최종키'] = np.where(df_godomall['자체옵션코드'] != '', df_godomall['자체옵션코드'], df_godomall[godo_name_col])
        df_godomall['merge_helper'] = df_godomall.groupby(['수취인 이름', '최종키']).cumcount()
        
        df_final['최종키'] = np.where(df_final['재고관리코드'] != '', df_final['재고관리코드'], df_final[ecount_name_col])
        df_final['merge_helper'] = df_final.groupby(['수령자명', '최종키']).cumcount()

        df_smartstore['최종키'] = np.where(df_smartstore['재고관리코드'] != '', df_smartstore['재고관리코드'], df_smartstore[smartstore_name_col])
        df_smartstore['merge_helper'] = df_smartstore.groupby(['수령자명', '최종키']).cumcount()
        
        # 6. 가격 정보 병합
        godo_price_map = df_godomall[['수취인 이름', '최종키', 'merge_helper', '수정될_금액_고도몰']]
        df_final = pd.merge(df_final, godo_price_map, left_on=['수령자명', '최종키', 'merge_helper'], right_on=['수취인 이름', '최종키', 'merge_helper'], how='left')
        
        smartstore_price_map = df_smartstore.rename(columns={'실결제금액': '수정될_금액_스토어'})
        smartstore_price_map = smartstore_price_map[['수령자명', '최종키', 'merge_helper', '수정될_금액_스토어']]
        df_final = pd.merge(df_final, smartstore_price_map, on=['수령자명', '최종키', 'merge_helper'], how='left')
                            
        # 7. 최종 금액 업데이트 및 임시 컬럼 정리
        df_final['실결제금액'] = np.where(df_final['쇼핑몰'] == '고도몰5', df_final['수정될_금액_고도몰'].fillna(df_final['실결제금액']), df_final['실결제금액'])
        df_final['실결제금액'] = np.where(df_final['쇼핑몰'] == '스마트스토어', df_final['수정될_금액_스토어'].fillna(df_final['실결제금액']), df_final['실결제금액'])
        df_final.drop(columns=['최종키', 'merge_helper', '수취인 이름', '수정될_금액_고도몰', '수정될_금액_스토어'], inplace=True, errors='ignore')
        
        # --- 나머지 처리 로직 (기존과 동일) ---
        warnings = []
        df_main_result = df_final[['재고관리코드', 'SKU상품명', '주문수량', '실결제금액', '쇼핑몰', '수령자명', 'original_order']]
        
        homonym_warnings = []
        name_groups = df_main_result.groupby('수령자명')['original_order'].apply(list)
        for name, orders in name_groups.items():
            if len(orders) > 1 and (max(orders) - min(orders) + 1) != len(orders):
                homonym_warnings.append(f"- [동명이인 의심] **{name}** 님의 주문이 떨어져서 입력되었습니다.")
        warnings.extend(homonym_warnings)

        df_quantity_summary = df_main_result.groupby('SKU상품명', as_index=False)['주문수량'].sum().rename(columns={'주문수량': '개수'})
        df_packing_list = df_main_result.sort_values(by='original_order')[['SKU상품명', '주문수량', '수령자명', '쇼핑몰']].copy()
        is_first_item = ~df_packing_list.duplicated(subset=['수령자명'], keep='first')
        df_packing_list['묶음번호'] = is_first_item.cumsum()
        df_packing_list['묶음번호'] = np.where(is_first_item, df_packing_list['묶음번호'], '')
        df_packing_list_final = df_packing_list[['묶음번호', 'SKU상품명', '주문수량', '수령자명', '쇼핑몰']]

        df_merged = pd.merge(df_main_result, df_master[['SKU코드', '과세여부', '입수량']], left_on='재고관리코드', right_on='SKU코드', how='left')
        
        unmastered = df_merged[df_merged['재고관리코드'].notna() & df_merged['SKU코드'].isna()]
        for _, row in unmastered.iterrows():
            warnings.append(f"- [미등록 상품] **{row['재고관리코드']}** / {row['SKU상품명']}")

        client_map = {'쿠팡': '쿠팡 주식회사', '고도몰5': '고래미자사몰_현금영수증(고도몰)', '스마트스토어': '스토어팜', '배민상회': '주식회사 우아한형제들(배민상회)', '이지웰몰': '주식회사 현대이지웰'}
        
        df_ecount_upload = pd.DataFrame()
        df_ecount_upload['일자'] = datetime.now().strftime("%Y%m%d")
        df_ecount_upload['거래처명'] = df_merged['쇼핑몰'].map(client_map).fillna(df_merged['쇼핑몰'])
        df_ecount_upload['출하창고'] = '고래미'
        df_ecount_upload['거래유형'] = np.where(df_merged['과세여부'] == '면세', 12, 11)
        df_ecount_upload['적요_전표'] = '오전/온라인'
        df_ecount_upload['품목코드'] = df_merged['재고관리코드']
        
        is_box_order = df_merged['SKU상품명'].str.contains("BOX", na=False)
        입수량 = pd.to_numeric(df_merged['입수량'], errors='coerce').fillna(1)
        base_quantity = np.where(is_box_order, df_merged['주문수량'] * 입수량, df_merged['주문수량'])
        is_3_pack = df_merged['SKU상품명'].str.contains("3개입|3개", na=False)
        final_quantity = np.where(is_3_pack, base_quantity * 3, base_quantity)
        df_ecount_upload['박스'] = np.where(is_box_order, df_merged['주문수량'], np.nan)
        df_ecount_upload['수량'] = final_quantity.astype(int)
        
        df_merged['실결제금액'] = pd.to_numeric(df_merged['실결제금액'], errors='coerce').fillna(0)
        공급가액 = np.where(df_merged['과세여부'] == '과세', df_merged['실결제금액'] / 1.1, df_merged['실결제금액'])
        df_ecount_upload['공급가액'] = 공급가액
        df_ecount_upload['부가세'] = df_merged['실결제금액'] - df_ecount_upload['공급가액']
        
        df_ecount_upload['쇼핑몰고객명'] = df_merged['수령자명']
        df_ecount_upload['original_order'] = df_merged['original_order']
        
        ecount_columns = ['일자', '순번', '거래처코드', '거래처명', '담당자', '출하창고', '거래유형', '통화', '환율', '적요_전표', '미수금', '총합계', '연결전표', '품목코드', '품목명', '규격', '박스', '수량', '단가', '외화금액', '공급가액', '부가세', '적요_품목', '생산전표생성', '시리얼/로트', '관리항목', '쇼핑몰고객명', 'original_order']
        for col in ecount_columns:
            if col not in df_ecount_upload:
                df_ecount_upload[col] = ''
        
        for col in ['공급가액', '부가세']:
            df_ecount_upload[col] = df_ecount_upload[col].round().astype('Int64')
        
        df_ecount_upload['거래유형'] = pd.to_numeric(df_ecount_upload['거래유형'])
        
        sort_order = ['고래미자사몰_현금영수증(고도몰)', '스토어팜', '쿠팡 주식회사', '주식회사 우아한형제들(배민상회)', '주식회사 현대이지웰']
        df_ecount_upload['거래처명_sort'] = pd.Categorical(df_ecount_upload['거래처명'], categories=sort_order, ordered=True)
        
        df_ecount_upload = df_ecount_upload.sort_values(by=['거래처명_sort', '거래유형', 'original_order'], ascending=[True, True, True]).drop(columns=['거래처명_sort', 'original_order'])
        
        df_ecount_upload = df_ecount_upload[ecount_columns[:-1]]

        return df_main_result.drop(columns=['original_order']), df_quantity_summary, df_packing_list_final, df_ecount_upload, True, "모든 파일 처리가 성공적으로 완료되었습니다.", warnings

    except Exception as e:
        import traceback
        st.error(f"처리 중 심각한 오류가 발생했습니다: {e}")
        st.error(traceback.format_exc())
        return None, None, None, None, False, f"오류가 발생했습니다. 파일을 다시 확인하거나 관리자에게 문의하세요.", []

# --------------------------------------------------------------------------
# Streamlit 앱 UI 구성 (이하 동일)
# --------------------------------------------------------------------------
st.set_page_config(page_title="주문 처리 자동화 v.Final-Masterpiece", layout="wide")
st.title("📑 주문 처리 자동화 (v.Final-Masterpiece)")
st.info("💡 3개의 주문 관련 파일을 업로드하면, 금액 보정, 물류, ERP(이카운트)용 데이터가 한 번에 생성됩니다.")
st.write("---")

st.header("1. 원본 엑셀 파일 3개 업로드")
col1, col2, col3 = st.columns(3)
with col1:
    file1 = st.file_uploader("1️⃣ 스마트스토어 (금액확인용)", type=['xlsx', 'xls', 'csv'])
with col2:
    file2 = st.file_uploader("2️⃣ 이카운트 다운로드 (주문목록)", type=['xlsx', 'xls', 'csv'])
with col3:
    file3 = st.file_uploader("3️⃣ 고도몰 (금액확인용)", type=['xlsx', 'xls', 'csv'])

st.write("---")
st.header("2. 처리 결과 확인 및 다운로드")
if st.button("🚀 모든 데이터 처리 및 파일 생성 실행"):
    if file1 and file2 and file3:
        try:
            df_master = load_local_master_data("master_data.csv")
            
            with st.spinner('모든 파일을 읽고 데이터를 처리하며 엑셀 서식을 적용 중입니다...'):
                df_main, df_qty, df_pack, df_ecount, success, message, warnings = process_all_files(file1, file2, file3, df_master)
            
            if success:
                st.success(message)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                if warnings:
                    st.warning("⚠️ 확인 필요 항목")
                    with st.expander("자세한 목록 보기..."):
                        st.info("금액 보정 실패, 미등록 상품, 동명이인 의심, 고도몰 금액 불일치 등의 데이터입니다. 원본 파일을 확인해주세요.")
                        for warning_message in warnings:
                            st.markdown(warning_message)
                
                tab_erp, tab_pack, tab_qty, tab_main = st.tabs(["🏢 **이카운트 업로드용**", "📋 포장 리스트", "📦 출고수량 요약", "✅ 최종 보정 리스트"])
                
                with tab_erp:
                    st.dataframe(df_ecount.astype(str))
                    st.download_button("📥 다운로드", to_excel_formatted(df_ecount, format_type='ecount_upload'), f"이카운트_업로드용_{timestamp}.xlsx")

                with tab_pack:
                    st.dataframe(df_pack)
                    st.download_button("📥 다운로드", to_excel_formatted(df_pack, format_type='packing_list'), f"물류팀_전달용_포장리스트_{timestamp}.xlsx")

                with tab_qty:
                    st.dataframe(df_qty)
                    st.download_button("📥 다운로드", to_excel_formatted(df_qty, format_type='quantity_summary'), f"물류팀_전달용_출고수량_{timestamp}.xlsx")
                
                with tab_main:
                    st.dataframe(df_main)
                    st.download_button("📥 다운로드", to_excel_formatted(df_main), f"최종_실결제금액_보정완료_{timestamp}.xlsx")
            else:
                # 오류 메시지는 process_all_files 함수 내에서 st.error()로 이미 표시됨
                pass
        
        except FileNotFoundError:
            st.error("🚨 치명적 오류: `master_data.csv` 파일을 찾을 수 없습니다! `app.py`와 동일한 폴더에 파일이 있는지 반드시 확인해주세요.")
        except Exception as e:
            st.error(f"🚨 상품 마스터 파일을 읽는 중 예상치 못한 오류가 발생했습니다: {e}")

    else:
        st.warning("⚠️ 3개의 엑셀 파일을 모두 업로드해야 실행할 수 있습니다.")
